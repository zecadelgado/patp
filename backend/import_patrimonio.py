"""
Módulo de Importação em Massa de Patrimônios
Permite importar milhares de patrimônios via planilha Excel/CSV
com criação automática de fornecedores, setores e categorias.

Versão: 2.4
Data: 19/11/2025
"""

import csv
import datetime
import os
from dataclasses import dataclass
from decimal import Decimal
from typing import Dict, List, Optional, Tuple

import openpyxl
from PySide6.QtCore import QObject, Signal
from PySide6.QtWidgets import QMessageBox

from database_manager import DatabaseManager
from validators import validar_cnpj, validar_email, validar_telefone


@dataclass
class PatrimonioImportRow:
    """Representa uma linha da planilha de importação"""
    linha: int
    nome: str
    descricao: Optional[str]
    numero_serie: Optional[str]
    data_aquisicao: datetime.date
    valor_compra: Decimal
    quantidade: int
    numero_nota: Optional[str]
    estado_conservacao: str
    categoria: str
    fornecedor_nome: str
    fornecedor_cnpj: Optional[str]
    fornecedor_telefone: Optional[str]
    fornecedor_email: Optional[str]
    setor_local: str
    status: str
    
    # Campos opcionais do fornecedor
    fornecedor_inscricao: Optional[str] = None
    fornecedor_contato: Optional[str] = None
    fornecedor_observacoes: Optional[str] = None


@dataclass
class ImportResult:
    """Resultado da importação"""
    sucesso: bool
    total_linhas: int
    importados: int
    erros: List[str]
    fornecedores_criados: int
    setores_criados: int
    categorias_criadas: int


class PatrimonioImporter(QObject):
    """Classe responsável pela importação em massa de patrimônios"""
    
    # Sinais para atualizar UI
    progress_updated = Signal(int, int)  # (atual, total)
    status_updated = Signal(str)
    
    # Colunas obrigatórias da planilha
    REQUIRED_COLUMNS = [
        "nome",
        "data_aquisicao",
        "valor_compra",
        "categoria",
        "fornecedor_nome",
        "setor_local",
        "status"
    ]
    
    # Colunas opcionais
    OPTIONAL_COLUMNS = [
        "descricao",
        "numero_serie",
        "quantidade",
        "numero_nota",
        "estado_conservacao",
        "fornecedor_cnpj",
        "fornecedor_telefone",
        "fornecedor_email",
        "fornecedor_inscricao",
        "fornecedor_contato",
        "fornecedor_observacoes"
    ]
    
    # Status válidos
    VALID_STATUS = ["ativo", "baixado", "em_manutencao", "desaparecido"]
    
    # Estados de conservação válidos
    VALID_ESTADOS = ["novo", "bom", "regular", "ruim"]
    
    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager = db_manager
        self._fornecedores_cache: Dict[str, int] = {}
        self._setores_cache: Dict[str, int] = {}
        self._categorias_cache: Dict[str, int] = {}
    
    def validar_arquivo(self, filepath: str) -> Tuple[bool, str]:
        """
        Valida se o arquivo existe e tem extensão válida
        
        Returns:
            (sucesso, mensagem)
        """
        if not os.path.exists(filepath):
            return False, "Arquivo não encontrado."
        
        ext = os.path.splitext(filepath)[1].lower()
        if ext not in ['.xlsx', '.xls', '.csv']:
            return False, "Formato inválido. Use .xlsx, .xls ou .csv"
        
        return True, "OK"
    
    def ler_planilha(self, filepath: str) -> Tuple[bool, List[Dict], str]:
        """
        Lê a planilha e retorna os dados
        
        Returns:
            (sucesso, dados, mensagem_erro)
        """
        ext = os.path.splitext(filepath)[1].lower()
        
        try:
            if ext == '.csv':
                return self._ler_csv(filepath)
            else:
                return self._ler_excel(filepath)
        except Exception as e:
            return False, [], f"Erro ao ler arquivo: {str(e)}"
    
    def _ler_excel(self, filepath: str) -> Tuple[bool, List[Dict], str]:
        """Lê arquivo Excel"""
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Ler cabeçalho
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
            
            if not headers:
                return False, [], "Planilha vazia ou sem cabeçalho."
            
            # Validar colunas obrigatórias
            missing = [col for col in self.REQUIRED_COLUMNS if col not in headers]
            if missing:
                return False, [], f"Colunas obrigatórias faltando: {', '.join(missing)}"
            
            # Ler dados
            dados = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Pular linhas vazias
                    continue
                
                row_dict = {"_linha": row_idx}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = value
                
                dados.append(row_dict)
            
            if not dados:
                return False, [], "Nenhum dado encontrado na planilha."
            
            return True, dados, "OK"
        
        except Exception as e:
            return False, [], f"Erro ao ler Excel: {str(e)}"
    
    def _ler_csv(self, filepath: str) -> Tuple[bool, List[Dict], str]:
        """Lê arquivo CSV"""
        try:
            dados = []
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                
                # Normalizar cabeçalhos
                fieldnames = [field.strip().lower() for field in reader.fieldnames]
                
                # Validar colunas obrigatórias
                missing = [col for col in self.REQUIRED_COLUMNS if col not in fieldnames]
                if missing:
                    return False, [], f"Colunas obrigatórias faltando: {', '.join(missing)}"
                
                # Ler dados
                for row_idx, row in enumerate(reader, start=2):
                    if not any(row.values()):
                        continue
                    
                    row_dict = {"_linha": row_idx}
                    for old_key, value in row.items():
                        new_key = old_key.strip().lower()
                        row_dict[new_key] = value
                    
                    dados.append(row_dict)
            
            if not dados:
                return False, [], "Nenhum dado encontrado no CSV."
            
            return True, dados, "OK"
        
        except Exception as e:
            return False, [], f"Erro ao ler CSV: {str(e)}"
    
    def validar_dados(self, dados: List[Dict]) -> Tuple[bool, List[PatrimonioImportRow], List[str]]:
        """
        Valida os dados da planilha
        
        Returns:
            (sucesso, linhas_validas, erros)
        """
        linhas_validas = []
        erros = []
        
        for row in dados:
            linha_num = row.get("_linha", 0)
            
            try:
                # Validar e converter dados
                patrimonio = self._validar_linha(row)
                linhas_validas.append(patrimonio)
            
            except ValueError as e:
                erros.append(f"Linha {linha_num}: {str(e)}")
            except Exception as e:
                erros.append(f"Linha {linha_num}: Erro inesperado - {str(e)}")
        
        sucesso = len(erros) == 0
        return sucesso, linhas_validas, erros
    
    def _validar_linha(self, row: Dict) -> PatrimonioImportRow:
        """Valida e converte uma linha da planilha"""
        linha_num = row.get("_linha", 0)
        
        # Nome (obrigatório)
        nome = str(row.get("nome", "")).strip()
        if not nome:
            raise ValueError("Nome é obrigatório")
        
        # Data de aquisição (obrigatório)
        data_aquisicao = self._parse_data(row.get("data_aquisicao"))
        if not data_aquisicao:
            raise ValueError("Data de aquisição inválida ou ausente")
        
        # Valor de compra (obrigatório)
        valor_compra = self._parse_decimal(row.get("valor_compra"))
        if valor_compra is None or valor_compra < 0:
            raise ValueError("Valor de compra inválido ou ausente")
        
        # Quantidade (opcional, padrão 1)
        quantidade = self._parse_int(row.get("quantidade", 1))
        if quantidade <= 0:
            raise ValueError("Quantidade deve ser maior que zero")
        
        # Categoria (obrigatório)
        categoria = str(row.get("categoria", "")).strip()
        if not categoria:
            raise ValueError("Categoria é obrigatória")
        
        # Fornecedor nome (obrigatório)
        fornecedor_nome = str(row.get("fornecedor_nome", "")).strip()
        if not fornecedor_nome:
            raise ValueError("Nome do fornecedor é obrigatório")
        
        # Fornecedor CNPJ (opcional, mas validar se fornecido)
        fornecedor_cnpj = str(row.get("fornecedor_cnpj", "")).strip() or None
        if fornecedor_cnpj and not validar_cnpj(fornecedor_cnpj):
            raise ValueError(f"CNPJ inválido: {fornecedor_cnpj}")
        
        # Fornecedor telefone (opcional, mas validar se fornecido)
        fornecedor_telefone = str(row.get("fornecedor_telefone", "")).strip() or None
        if fornecedor_telefone and not validar_telefone(fornecedor_telefone):
            raise ValueError(f"Telefone inválido: {fornecedor_telefone}")
        
        # Fornecedor email (opcional, mas validar se fornecido)
        fornecedor_email = str(row.get("fornecedor_email", "")).strip() or None
        if fornecedor_email and not validar_email(fornecedor_email):
            raise ValueError(f"Email inválido: {fornecedor_email}")
        
        # Setor/Local (obrigatório)
        setor_local = str(row.get("setor_local", "")).strip()
        if not setor_local:
            raise ValueError("Setor/Local é obrigatório")
        
        # Status (obrigatório e deve ser válido)
        status = str(row.get("status", "ativo")).strip().lower()
        if status not in self.VALID_STATUS:
            raise ValueError(f"Status inválido: {status}. Use: {', '.join(self.VALID_STATUS)}")
        
        # Estado de conservação (opcional, padrão "bom")
        estado_conservacao = str(row.get("estado_conservacao", "bom")).strip().lower()
        if estado_conservacao not in self.VALID_ESTADOS:
            raise ValueError(f"Estado inválido: {estado_conservacao}. Use: {', '.join(self.VALID_ESTADOS)}")
        
        # Campos opcionais
        descricao = str(row.get("descricao", "")).strip() or None
        numero_serie = str(row.get("numero_serie", "")).strip() or None
        numero_nota = str(row.get("numero_nota", "")).strip() or None
        fornecedor_inscricao = str(row.get("fornecedor_inscricao", "")).strip() or None
        fornecedor_contato = str(row.get("fornecedor_contato", "")).strip() or None
        fornecedor_observacoes = str(row.get("fornecedor_observacoes", "")).strip() or None
        
        return PatrimonioImportRow(
            linha=linha_num,
            nome=nome,
            descricao=descricao,
            numero_serie=numero_serie,
            data_aquisicao=data_aquisicao,
            valor_compra=valor_compra,
            quantidade=quantidade,
            numero_nota=numero_nota,
            estado_conservacao=estado_conservacao,
            categoria=categoria,
            fornecedor_nome=fornecedor_nome,
            fornecedor_cnpj=fornecedor_cnpj,
            fornecedor_telefone=fornecedor_telefone,
            fornecedor_email=fornecedor_email,
            fornecedor_inscricao=fornecedor_inscricao,
            fornecedor_contato=fornecedor_contato,
            fornecedor_observacoes=fornecedor_observacoes,
            setor_local=setor_local,
            status=status
        )
    
    def _parse_data(self, value) -> Optional[datetime.date]:
        """Converte valor para data"""
        if value is None or value == "":
            return None
        
        if isinstance(value, datetime.datetime):
            return value.date()
        
        if isinstance(value, datetime.date):
            return value
        
        # Tentar parsear string
        value_str = str(value).strip()
        
        # Formatos aceitos: DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY
        formats = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]
        
        for fmt in formats:
            try:
                return datetime.datetime.strptime(value_str, fmt).date()
            except:
                continue
        
        return None
    
    def _parse_decimal(self, value) -> Optional[Decimal]:
        """Converte valor para Decimal"""
        if value is None or value == "":
            return None
        
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        # Limpar string (remover R$, espaços, etc)
        value_str = str(value).strip()
        value_str = value_str.replace("R$", "").replace(" ", "")
        value_str = value_str.replace(".", "").replace(",", ".")  # Formato BR -> US
        
        try:
            return Decimal(value_str)
        except:
            return None
    
    def _parse_int(self, value) -> int:
        """Converte valor para int"""
        if isinstance(value, int):
            return value
        
        if isinstance(value, float):
            return int(value)
        
        try:
            return int(str(value).strip())
        except:
            return 1
    
    def importar(self, linhas: List[PatrimonioImportRow]) -> ImportResult:
        """
        Importa os patrimônios para o banco de dados
        
        Returns:
            ImportResult com estatísticas da importação
        """
        total = len(linhas)
        importados = 0
        erros = []
        fornecedores_criados = 0
        setores_criados = 0
        categorias_criadas = 0
        
        # Carregar caches
        self._carregar_caches()
        
        for idx, linha in enumerate(linhas, start=1):
            try:
                self.status_updated.emit(f"Importando linha {idx} de {total}...")
                self.progress_updated.emit(idx, total)
                
                # 1. Garantir que fornecedor existe
                fornecedor_id, criado = self._garantir_fornecedor(linha)
                if criado:
                    fornecedores_criados += 1
                
                # 2. Garantir que setor existe
                setor_id, criado = self._garantir_setor(linha.setor_local)
                if criado:
                    setores_criados += 1
                
                # 3. Garantir que categoria existe
                categoria_id, criado = self._garantir_categoria(linha.categoria)
                if criado:
                    categorias_criadas += 1
                
                # 4. Criar patrimônio
                self._criar_patrimonio(linha, fornecedor_id, setor_id, categoria_id)
                
                importados += 1
            
            except Exception as e:
                erros.append(f"Linha {linha.linha}: {str(e)}")
        
        sucesso = len(erros) == 0
        
        return ImportResult(
            sucesso=sucesso,
            total_linhas=total,
            importados=importados,
            erros=erros,
            fornecedores_criados=fornecedores_criados,
            setores_criados=setores_criados,
            categorias_criadas=categorias_criadas
        )
    
    def _carregar_caches(self):
        """Carrega caches de fornecedores, setores e categorias"""
        # Fornecedores
        try:
            fornecedores = self.db_manager.fetch_all(
                "SELECT id_fornecedor, nome_fornecedor, cnpj FROM fornecedores"
            )
            for f in fornecedores:
                nome = str(f['nome_fornecedor']).strip().lower()
                self._fornecedores_cache[nome] = f['id_fornecedor']
                
                if f.get('cnpj'):
                    cnpj = str(f['cnpj']).strip()
                    self._fornecedores_cache[f"cnpj:{cnpj}"] = f['id_fornecedor']
        except:
            pass
        
        # Setores
        try:
            setores = self.db_manager.fetch_all(
                "SELECT id_setor_local, nome_setor_local FROM setores_locais"
            )
            for s in setores:
                nome = str(s['nome_setor_local']).strip().lower()
                self._setores_cache[nome] = s['id_setor_local']
        except:
            pass
        
        # Categorias
        try:
            categorias = self.db_manager.fetch_all(
                "SELECT id_categoria, nome_categoria FROM categorias"
            )
            for c in categorias:
                nome = str(c['nome_categoria']).strip().lower()
                self._categorias_cache[nome] = c['id_categoria']
        except:
            pass
    
    def _garantir_fornecedor(self, linha: PatrimonioImportRow) -> Tuple[int, bool]:
        """
        Garante que o fornecedor existe, criando se necessário
        
        Returns:
            (id_fornecedor, foi_criado)
        """
        nome_lower = linha.fornecedor_nome.strip().lower()
        
        # Verificar cache por nome
        if nome_lower in self._fornecedores_cache:
            return self._fornecedores_cache[nome_lower], False
        
        # Verificar cache por CNPJ
        if linha.fornecedor_cnpj:
            cnpj_key = f"cnpj:{linha.fornecedor_cnpj}"
            if cnpj_key in self._fornecedores_cache:
                return self._fornecedores_cache[cnpj_key], False
        
        # Criar fornecedor
        dados = {
            "nome_fornecedor": linha.fornecedor_nome,
            "cnpj": linha.fornecedor_cnpj,
            "telefone": linha.fornecedor_telefone,
            "email": linha.fornecedor_email,
            "inscricao_estadual": linha.fornecedor_inscricao,
            "contato": linha.fornecedor_contato,
            "observacoes": linha.fornecedor_observacoes
        }
        
        # Filtrar None
        dados = {k: v for k, v in dados.items() if v is not None}
        
        # Inserir no banco
        cursor = self.db_manager.connection.cursor()
        try:
            columns = ", ".join(f"`{k}`" for k in dados.keys())
            placeholders = ", ".join(["%s"] * len(dados))
            sql = f"INSERT INTO fornecedores ({columns}) VALUES ({placeholders})"
            cursor.execute(sql, tuple(dados.values()))
            self.db_manager.connection.commit()
            
            fornecedor_id = cursor.lastrowid
            
            # Atualizar cache
            self._fornecedores_cache[nome_lower] = fornecedor_id
            if linha.fornecedor_cnpj:
                self._fornecedores_cache[f"cnpj:{linha.fornecedor_cnpj}"] = fornecedor_id
            
            return fornecedor_id, True
        
        finally:
            cursor.close()
    
    def _garantir_setor(self, nome_setor: str) -> Tuple[int, bool]:
        """
        Garante que o setor existe, criando se necessário
        
        Returns:
            (id_setor, foi_criado)
        """
        nome_lower = nome_setor.strip().lower()
        
        # Verificar cache
        if nome_lower in self._setores_cache:
            return self._setores_cache[nome_lower], False
        
        # Criar setor
        cursor = self.db_manager.connection.cursor()
        try:
            sql = "INSERT INTO setores_locais (nome_setor_local) VALUES (%s)"
            cursor.execute(sql, (nome_setor,))
            self.db_manager.connection.commit()
            
            setor_id = cursor.lastrowid
            
            # Atualizar cache
            self._setores_cache[nome_lower] = setor_id
            
            return setor_id, True
        
        finally:
            cursor.close()
    
    def _garantir_categoria(self, nome_categoria: str) -> Tuple[int, bool]:
        """
        Garante que a categoria existe, criando se necessário
        
        Returns:
            (id_categoria, foi_criado)
        """
        nome_lower = nome_categoria.strip().lower()
        
        # Verificar cache
        if nome_lower in self._categorias_cache:
            return self._categorias_cache[nome_lower], False
        
        # Criar categoria
        cursor = self.db_manager.connection.cursor()
        try:
            sql = "INSERT INTO categorias (nome_categoria) VALUES (%s)"
            cursor.execute(sql, (nome_categoria,))
            self.db_manager.connection.commit()
            
            categoria_id = cursor.lastrowid
            
            # Atualizar cache
            self._categorias_cache[nome_lower] = categoria_id
            
            return categoria_id, True
        
        finally:
            cursor.close()
    
    def _criar_patrimonio(
        self, 
        linha: PatrimonioImportRow,
        fornecedor_id: int,
        setor_id: int,
        categoria_id: int
    ):
        """Cria o patrimônio no banco de dados"""
        dados = {
            "nome": linha.nome,
            "descricao": linha.descricao,
            "numero_serie": linha.numero_serie,
            "data_aquisicao": linha.data_aquisicao,
            "valor_compra": float(linha.valor_compra),
            "quantidade": linha.quantidade,
            "numero_nota": linha.numero_nota,
            "estado_conservacao": linha.estado_conservacao,
            "id_categoria": categoria_id,
            "id_fornecedor": fornecedor_id,
            "id_setor_local": setor_id,
            "status": linha.status
        }
        
        # Filtrar None
        dados = {k: v for k, v in dados.items() if v is not None}
        
        # Inserir no banco
        cursor = self.db_manager.connection.cursor()
        try:
            columns = ", ".join(f"`{k}`" for k in dados.keys())
            placeholders = ", ".join(["%s"] * len(dados))
            sql = f"INSERT INTO patrimonios ({columns}) VALUES ({placeholders})"
            cursor.execute(sql, tuple(dados.values()))
            self.db_manager.connection.commit()
        
        finally:
            cursor.close()
