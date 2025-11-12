from __future__ import annotations

from typing import Any, Dict, List, Optional, Sequence

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QFileDialog,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QWidget,
)

from database_manager import DatabaseManager


class RelatoriosController:
    """Popular os relatórios agregados do sistema."""

    def __init__(self, widget: QWidget, db_manager: DatabaseManager) -> None:
        self.widget = widget
        self.db_manager = db_manager

        self.tbl_por_setor: Optional[QTableWidget] = self.widget.findChild(QTableWidget, "tbl_por_setor")
        self.tbl_por_categoria: Optional[QTableWidget] = self.widget.findChild(QTableWidget, "tbl_por_categoria")
        self.tbl_manutencoes: Optional[QTableWidget] = self.widget.findChild(QTableWidget, "tbl_manutencoes")

        btn = self.widget.findChild(QPushButton, "btn_gerar_relatorio")
        if btn is None:
            btn = QPushButton("Gerar Relatório (.xlsx)", self.widget)
            btn.setObjectName("btn_gerar_relatorio")
            lay = self.widget.layout()
            if lay is not None:
                lay.addWidget(btn)
        btn.clicked.connect(self.gerar_relatorio_excel)

        self.refresh()

                                                                          
    def refresh(self) -> None:
        self._carregar_por_setor()
        self._carregar_por_categoria()
        self._carregar_manutencoes()

                                                                          
    def _carregar_por_setor(self) -> None:
        try:
            rows = self.db_manager.relatorio_por_setor()
        except Exception as exc:                                       
            print(f"[Relatórios] Falha ao carregar relatório por setor: {exc}")
            rows = []
        linhas = [
            [
                row.get("setor") or "-",
                row.get("localizacao") or "-",
                str(row.get("quantidade") or 0),
                self._format_currency(row.get("valor_total")),
            ]
            for row in rows
        ]
        self._popular_tabela(self.tbl_por_setor, linhas)

    def _carregar_por_categoria(self) -> None:
        try:
            rows = self.db_manager.relatorio_por_categoria()
        except Exception as exc:                                       
            print(f"[Relatórios] Falha ao carregar relatório por categoria: {exc}")
            rows = []
        linhas = [
            [
                row.get("categoria") or "-",
                str(row.get("quantidade") or 0),
                self._format_currency(row.get("depreciacao_acumulada")),
            ]
            for row in rows
        ]
        self._popular_tabela(self.tbl_por_categoria, linhas)

    def _carregar_manutencoes(self) -> None:
        try:
            rows = self.db_manager.relatorio_manutencoes()
        except Exception as exc:                                       
            print(f"[Relatórios] Falha ao carregar relatório de manutenções: {exc}")
            rows = []
        linhas = [
            [
                self._format_date(row.get("data_inicio")),
                row.get("nome_patrimonio") or "-",
                row.get("tipo") or "-",
                row.get("responsavel") or "-",
                self._format_currency(row.get("custo")),
            ]
            for row in rows
        ]
        self._popular_tabela(self.tbl_manutencoes, linhas)

                                                                          
    def _popular_tabela(self, tabela: Optional[QTableWidget], linhas: Sequence[Sequence[str]]) -> None:
        if not tabela:
            return
        tabela.setRowCount(len(linhas))
        for row_index, linha in enumerate(linhas):
            for col_index, valor in enumerate(linha):
                item = QTableWidgetItem(valor)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if valor.startswith("R$"):
                    item.setTextAlignment(int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter))
                tabela.setItem(row_index, col_index, item)
        tabela.resizeColumnsToContents()

    @staticmethod
    def _format_currency(value: Optional[object]) -> str:
        try:
            return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (TypeError, ValueError):
            return "R$ 0,00"

    @staticmethod
    def _format_date(value: Optional[object]) -> str:
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        return str(value or "-")

    def gerar_relatorio_excel(self) -> None:
        """
        Gera um arquivo .xlsx com todos os patrimônios, 1 por linha, e inclui
        uma linha final com o somatório de valores.
        - Usa DatabaseManager.list_patrimonios(None) para obter os dados.
        - Se encontrar registros legados com quantidade>1, “explode” em múltiplas linhas
          para cumprir o requisito de individualização visual no relatório.
          (Obs: o ID será repetido nesses casos legados, pois o banco tem 1 linha para N unidades.)
        """
        try:
            rows = self.db_manager.list_patrimonios(None) or []
        except Exception as exc:
            QMessageBox.critical(self.widget, "Relatórios", f"Falha ao consultar patrimônios.\n{exc}")
            return

        if not rows:
            QMessageBox.information(self.widget, "Relatórios", "Não há patrimônios para exportar.")
            return

        columns = [
            ("id_patrimonio", "ID"),
            ("nome_patrimonio", "Nome"),
            ("descricao", "Descrição"),
            ("numero_serie", "Número de Série"),
            ("status", "Situação"),
            ("estado_conservacao", "Estado de Conservação"),
            ("data_aquisicao", "Data de Aquisição"),
            ("id_categoria", "ID Categoria"),
            ("nome_categoria", "Categoria"),
            ("id_fornecedor", "ID Fornecedor"),
            ("nome_fornecedor", "Fornecedor"),
            ("id_setor_local", "ID Setor/Local"),
            ("nome_setor_local", "Setor/Local"),
            ("centros_custo", "Centros de Custo"),
            ("numero_nota", "Número da Nota"),
            ("valor_compra", "Valor de Compra"),
            ("valor_atual", "Valor Atual"),
            ("quantidade", "Quantidade (registro)"),
        ]

        exploded: List[Dict[str, Any]] = []
        for r in rows:
            q = 1
            try:
                q = int(r.get("quantidade") or 1)
            except Exception:
                q = 1
            q = max(1, q)
            for _ in range(q):
                exploded.append(r)

        filename, _ = QFileDialog.getSaveFileName(
            self.widget,
            "Salvar Relatório de Patrimônios",
            "relatorio_patrimonios.xlsx",
            "Planilhas Excel (*.xlsx)",
        )
        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception as exc:
            QMessageBox.critical(
                self.widget,
                "Relatórios",
                "Biblioteca 'openpyxl' não encontrada.\nInstale-a e tente novamente.\n"
                f"{exc}",
            )
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Patrimônios"

        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        currency = NamedStyle(name="brl")
        currency.number_format = '"R$" #,##0.00'

        for col_idx, (_, label) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total_val = 0.0
        for row_idx, item in enumerate(exploded, start=2):
            for col_idx, (key, _) in enumerate(columns, start=1):
                val = item.get(key)
                if key in ("valor_atual", "valor_compra"):
                    try:
                        valf = float(val or 0.0)
                    except Exception:
                        valf = 0.0
                    ws.cell(row=row_idx, column=col_idx, value=valf).style = currency
                elif key == "data_aquisicao" and hasattr(val, "strftime"):
                    ws.cell(row=row_idx, column=col_idx, value=val.strftime("%Y-%m-%d"))
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)

            try:
                v = float(
                    item.get("valor_atual")
                    if item.get("valor_atual") is not None
                    else item.get("valor_compra")
                    or 0.0
                )
            except Exception:
                v = 0.0
            total_val += v

        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        last_row = len(exploded) + 2
        ws.cell(row=last_row, column=len(columns) - 2, value="TOTAL").font = header_font
        tv_cell = ws.cell(row=last_row, column=len(columns) - 1, value=total_val)
        tv_cell.style = currency
        tv_cell.font = header_font

        try:
            wb.save(filename)
        except Exception as exc:
            QMessageBox.critical(self.widget, "Relatórios", f"Falha ao salvar Excel.\n{exc}")
            return

        QMessageBox.information(
            self.widget,
            "Relatórios",
            f"Relatório gerado com sucesso:\n{filename}",
        )
